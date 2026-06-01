"""
Spreadsheet Export Service — Generates Excel workbooks from extracted PRF data.
Creates a structured spreadsheet with all extracted fields from processed documents.
"""
from __future__ import annotations
import io
from datetime import datetime, timezone
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Column definitions for PRF spreadsheet
PRF_COLUMNS = [
    ("Document ID", "id", 15),
    ("Filename", "original_filename", 25),
    ("OCR Status", "ocr_status", 12),
    ("Confidence", "ocr_confidence_avg", 10),
    ("Method", "method_used", 12),
    ("Patient Name", "patient_name", 25),
    ("SA ID Number", "patient_id_number", 15),
    ("Date of Birth", "patient_dob", 14),
    ("Gender", "gender", 8),
    ("Phone", "patient_phone", 15),
    ("Medical Scheme", "medical_scheme", 22),
    ("Member Number", "member_number", 15),
    ("Dependent Code", "dependent_code", 12),
    ("Incident Date", "incident_date", 14),
    ("Incident Time", "incident_time", 12),
    ("Incident Location", "incident_location", 30),
    ("Chief Complaint", "chief_complaint", 30),
    ("Clinical Notes", "clinical_notes", 40),
    ("Mechanism of Injury", "mechanism_of_injury", 20),
    ("ICD-10 Codes", "icd10_codes", 18),
    ("Treating Provider", "treating_provider", 20),
    ("Practice Number", "provider_practice_number", 15),
    ("Crew Members", "crew_members", 25),
    ("Vehicle Reg", "vehicle_registration", 14),
    ("Vehicle Callsign", "vehicle_callsign", 14),
    ("Dispatch Time", "dispatch_time", 12),
    ("On Scene Time", "on_scene_time", 12),
    ("Transport Time", "transport_time", 12),
    ("Hospital Arrival", "hospital_arrival_time", 14),
    ("Receiving Facility", "receiving_facility", 22),
    ("Pre-Auth Number", "preauth_number", 15),
    ("BP Sys/Dia", "bp", 12),
    ("Heart Rate", "heart_rate", 10),
    ("Resp Rate", "respiratory_rate", 10),
    ("SpO2", "spo2", 8),
    ("GCS", "gcs", 6),
    ("Temperature", "temperature", 10),
    ("Blood Glucose", "blood_glucose", 12),
    ("Pain Scale", "pain_scale", 10),
    ("Medications", "medications_administered", 30),
    ("Procedures", "procedures_performed", 30),
    ("Patient Sig", "patient_signature_present", 10),
    ("Provider Sig", "provider_signature_present", 10),
    ("Total Amount", "total_amount", 12),
    ("Tariff Codes", "tariff_codes", 20),
    ("Processed At", "created_at", 20),
]

# Header styles
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Section header fills
SECTION_FILLS = {
    "document": PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"),
    "patient": PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid"),
    "incident": PatternFill(start_color="117A65", end_color="117A65", fill_type="solid"),
    "clinical": PatternFill(start_color="6C3483", end_color="6C3483", fill_type="solid"),
    "vitals": PatternFill(start_color="B7950B", end_color="B7950B", fill_type="solid"),
    "transport": PatternFill(start_color="A04000", end_color="A04000", fill_type="solid"),
    "billing": PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid"),
}

DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def _extract_field(doc_data: dict, extracted: dict, field_key: str) -> str:
    """Extract a field value from document data or extracted data."""
    # Check document-level fields first
    if field_key in doc_data:
        val = doc_data[field_key]
        if val is not None:
            return str(val)

    # Check extracted data
    if field_key in extracted:
        val = extracted[field_key]
        if isinstance(val, list):
            return ", ".join(str(v) for v in val)
        if isinstance(val, dict):
            return "; ".join(f"{k}={v}" for k, v in val.items() if v)
        if isinstance(val, bool):
            return "Yes" if val else "No"
        return str(val) if val is not None else ""

    # Handle vital signs (nested dict)
    if field_key in ("heart_rate", "respiratory_rate", "spo2", "gcs",
                      "temperature", "blood_glucose", "pain_scale"):
        vitals = extracted.get("vital_signs", {})
        if isinstance(vitals, dict):
            val = vitals.get(field_key)
            if val:
                return str(val)

    # Handle special compound fields
    if field_key == "bp":
        vitals = extracted.get("vital_signs", {})
        if isinstance(vitals, dict):
            sys_bp = vitals.get("bp_systolic", "")
            dia_bp = vitals.get("bp_diastolic", "")
            if sys_bp or dia_bp:
                return f"{sys_bp}/{dia_bp}"

    if field_key == "tariff_codes":
        items = extracted.get("line_items", [])
        if isinstance(items, list):
            codes = [str(i.get("tariff_code", "")) for i in items if isinstance(i, dict) and i.get("tariff_code")]
            return ", ".join(codes)

    if field_key == "method_used":
        return doc_data.get("method_used", "")

    return ""


def generate_prf_spreadsheet(documents: List[dict]) -> bytes:
    """
    Generate an Excel spreadsheet from a list of processed document dicts.
    Each document dict should contain: id, original_filename, ocr_status,
    ocr_confidence_avg, extracted_data, created_at, etc.

    Returns: Excel file as bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "PRF Extractions"

    # ── Write headers ──
    for col_idx, (label, _, width) in enumerate(PRF_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    # ── Write data rows ──
    for row_idx, doc in enumerate(documents, start=2):
        extracted = doc.get("extracted_data", {}) or {}

        for col_idx, (_, field_key, _) in enumerate(PRF_COLUMNS, start=1):
            value = _extract_field(doc, extracted, field_key)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # ── Summary sheet ──
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 30

    summary_data = [
        ("Report", "PRF Extraction Data"),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Total Documents", str(len(documents))),
        ("Completed", str(sum(1 for d in documents if d.get("ocr_status") == "completed"))),
        ("Failed", str(sum(1 for d in documents if d.get("ocr_status") == "failed"))),
        ("Pending Review", str(sum(1 for d in documents if d.get("needs_hitl_review")))),
    ]

    for row_idx, (label, value) in enumerate(summary_data, start=1):
        label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
        label_cell.font = Font(name="Calibri", bold=True, size=11)
        value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
        value_cell.font = Font(name="Calibri", size=11)

    # ── Save to bytes ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
